# ------------------------------------------------------------------------ #
# Title: Excel Write
# Description: This program writes the MRF data to an excel file

# ------------------------------------------------------------------------ #
from openpyxl import load_workbook
import datetime
import os
# import pickle
# from vendor_address_info import vendor_address

if __name__ == "__main__":
    raise Exception("This file is not meant to ran by itself")

# open work book
template_path = os.getcwd() + "\\Material Request Form_template.xlsx"
#template_path = os.getcwd() + r"\MRF_template - Copy.xlsx"
wb = load_workbook(template_path)
ws = wb['MRF']

# define date constant
d = datetime.datetime.today()


# create a function with item attribute inputs and write the input to the excel template
class XL:

    @staticmethod
    def write_item(item_list):
        """Writes Item data to the excel sheet.
        :param item_list: list passed from the new_item instantiation in the form of .mrf_list_format
        :return don't return anything just do it"""

        column_letters = ['A', 'C', 'K', 'S', 'AA', 'AO', 'AW', 'AY']
        # The first row of listing the items begins on 20
        # fills out the rows
        z = 14
        for item in item_list:
            z += 1
            for k, letter in enumerate(column_letters):
                ws[letter + str(z)] = item[k]

    @staticmethod
    def save_sheet(savename, savedir, vendor):
        """This function will save the new MRF to the appropriate directory.
        Save name is "MMDDYY_hour_minute-initials vendor"
        :param savename: name of file
        :param savedir: file directory where the excel file will be saved
        :return: String statement confirming the excel document has been saved, with which path its been saved to
        """

        # savename is mm/dd/yy__hour.minute--Initials Vendor name
        # savename = d.strftime('%m%d%y_%H_%M') + "-" + initials + ' ' + str(vendor)
        savepath = savedir + '\\' + savename + ' ' + str(vendor) + ".xlsx"
        wb.save(savepath)
        return print("File was saved to: " + savedir + "\nFilename: " + savename)

    # function to save the file with appropriate naming schema
    # todo: add initials variable to allow for multiple users initials
    # todo: create if/else which can navigate the file directory to save
    #       the file automatically to the correct month/year folder on the work computers

    @staticmethod
    # def write_mrf_details(initials, vendor, requestor, RMA, shipping_info,  cert_type, interval, cal_specs):
    def write_mrf_details(mrn, company, shipping_acx, rma, index, gage_id, cert_template, vendor):
        '''

        :param mrn: Material request number, entered by user
        :param company: Client company obtained from query
        :param shipping_info: shipping info obtained from shipping_accs.py
        :param index: index will be used to either write gid or 'multiple'
        :param rma: return material authorization number
        :param gage_id: gage id
        :param cert_template: certificate level.
        :param vendor: name of the vendor
        :return: none
        '''
        # material_request_num
        ws['K9'] = mrn
        # date requested, date required
        ws['AO11'] = d.strftime('%m/%d/%Y')
        # adds one month to the date
        ws['BA11'] = str((int(d.strftime('%m')) + 1)) + d.strftime('/%d/%Y')
        # requestor = User.full_name() -- use pickling to save the user info?
        ws['K29'] = 'William Ang'
        # company name
        ws['AO10'] = company
        # rma
        ws['AO9'] = 'RMA: ' + rma
        ws['AW6'] = rma
        # certificate template
        ws['K11'] = cert_template
        # shipping acx
        #for item in shipping_acx:

        ws['K12'] = "Carrier: " + shipping_acx[1] + "  Acx #: " + shipping_acx[0]
        # vendor address (E31 and down)
        # j = 31
        # ws['E30'] = "Ship to:"
        # for line in vendor_address(vendor):
        #
        #     ws['E' + str(j)] = line
        #     j += 1



        # vendor.get_address()
        # cert_type = Item.cert_type() unless otherwise specified
        # interval = Item.interval()
        # cal_specs = "cal to mfg specs" unless otherwise specified.
